from flask import Flask, render_template, jsonify, request
import pandas as pd
import os

app = Flask(__name__)

# Get the directory where this script is located
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, 'data', 'CEFR_Validation_Review_Neo4j.xlsx')
SHEET_NAME_SUMMARY = 'Summary'

# Load the data - simple caching mechanism
def load_data():
    # We want to read all sheets except Summary if needed, or iterate through them.
    # Based on the guide, "Review each level's sheet (A1, A2, B1, etc.)".
    # We should probably combine them for the app view, or provide a selector.
    # Let's combine all flagged items from A1, A2, B1, B2, C1, C2 sheets.
    xl = pd.ExcelFile(DATA_FILE)
    all_sheets = []
    
    # Load all sheets except Summary to catch levels like A1+, etc.
    excluded_sheets = [SHEET_NAME_SUMMARY]
    
    for sheet in xl.sheet_names:
        if sheet not in excluded_sheets:
            df = pd.read_excel(DATA_FILE, sheet_name=sheet)
            # Ensure we have the necessary columns.
            # Convert NaN to empty string/None for JSON dumping
            df = df.fillna('')
            # Add a 'Level' column for context
            df['Level'] = sheet
            all_sheets.append(df)
            
    if not all_sheets:
        return pd.DataFrame()
        
    full_df = pd.concat(all_sheets, ignore_index=True)
    return full_df

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/data')
def get_data():
    try:
        df = load_data()
        
        # Standardize keys
        records = []
        for _, row in df.iterrows():
            record = {
                'Level': row.get('Level', ''),
                'ID': row.get('ID', ''),
                'Descriptor': row.get('Descriptor Text', ''),
                'Category': row.get('Assigned Category', ''),
                'IsFlagged': row.get('Issue?', '') == 'YES',
                'IssueType': row.get('Issue Type', ''),
                'Evidence': row.get('Evidence', ''),
                'ExpertReview': row.get('Expert: Correct?', ''), # Check logic below if column differs
                'NewCategory': row.get('Expert: New Category', ''),
                'Confidence': row.get('Expert: Confidence', ''),
                'Notes': row.get('Expert: Notes', '')
            }
            # Fallback for Expert Review if my guess is wrong
            # Based on dump: 0:#, 1:ID, 2:Desc, 3:Assigned, 4:Issue?, 5:IssueType, 6:Evidence
            # 7:Expert: Correct? (Hypothesis), 8:Expert: New Category, 9:Expert: Confidence, 10:Expert: Notes
            # I'll rely on the dump I just read.
            
            records.append(record)
            
        return jsonify(records)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/save', methods=['POST'])
def save_data():
    try:
        data = request.json
        descriptor_id = data.get('ID')
        decision = data.get('Decision') 
        new_category = data.get('NewCategory') 
        confidence = data.get('Confidence') 
        notes = data.get('Notes') 
        level = data.get('Level') 
        
        if not descriptor_id or not level:
            return jsonify({'error': 'Missing ID or Level'}), 400

        # Load specific sheet
        from openpyxl import load_workbook
        wb = load_workbook(DATA_FILE)
        if level not in wb.sheetnames:
             return jsonify({'error': f'Sheet {level} not found'}), 404
             
        ws = wb[level]
        
        # Find headers to map columns dynamically
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers[val] = col
                
        # Target Columns
        col_decision = headers.get('Expert: Correct?')
        col_new_cat = headers.get('Expert: New Category')
        col_conf = headers.get('Expert: Confidence')
        col_notes = headers.get('Expert: Notes')
        col_id = headers.get('ID')
        
        if not col_id:
             return jsonify({'error': 'ID column not found'}), 500

        target_row = None
        for row in range(2, ws.max_row + 1): 
            cell_val = ws.cell(row=row, column=col_id).value
            if str(cell_val) == str(descriptor_id):
                target_row = row
                break
        
        if target_row:
            if decision is not None and col_decision: ws.cell(row=target_row, column=col_decision).value = decision
            if new_category is not None and col_new_cat: ws.cell(row=target_row, column=col_new_cat).value = new_category
            if confidence is not None and col_conf: ws.cell(row=target_row, column=col_conf).value = confidence
            if notes is not None and col_notes: ws.cell(row=target_row, column=col_notes).value = notes
            
            wb.save(DATA_FILE)
            return jsonify({'status': 'success'})
        else:
            return jsonify({'error': 'Descriptor not found in sheet'}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
